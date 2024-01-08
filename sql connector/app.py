from flask import Flask,request
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///todo.db'

db = SQLAlchemy(app)

class student(db.Model):
    roll_no = db.Column(db.Integer, primary_key=True,auto_increment=False)
    name = db.Column(db.String(100))
    percentage = db.Column(db.Integer)
    branch = db.Column(db.String(50))
    
@app.route('/get', methods= ['POST'])
def add_post():
    if request.method == 'POST':
        data = request.files['abdul']
        abdul =load_workbook(data)
        abdul1=abdul.active
        for i in abdul1.iter_rows(min_row=2,values_only = True):
            data1 = student(roll_no=i[0], name=i[1],percentage=i[2],branch=i[3])
            db.session.add(data1)
        db.session.commit()
    return "msg:""data retrieve"
with app.app_context():
    db.create_all()
if __name__ == "__main__":
     app.run(debug=True ,port=8000,use_reloader=False)
    